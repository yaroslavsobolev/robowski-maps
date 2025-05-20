from robowski.settings import *
import os
import pandas as pd
import openpyxl as xl

def request_outliers(run_name, destination_run, known_outliers_filename, manual_outliers_filename,
                     n_repetitions_of_manual_outliers=2):

    #load outliers
    df_manual_outliers = pd.read_csv(data_folder + run_name + f'results/outliers/{manual_outliers_filename}.csv')
    if known_outliers_filename:
        df_known_outliers = pd.read_csv(data_folder + run_name + f'results/outliers/{known_outliers_filename}.csv')
        df_all_outliers = pd.concat([df_known_outliers] + [df_manual_outliers] * n_repetitions_of_manual_outliers,
                                    ignore_index=True, sort=False)
    else:
        df_all_outliers = pd.concat([df_manual_outliers] * n_repetitions_of_manual_outliers,
                                    ignore_index=True, sort=False)

    # randomize the order
    df_all_outliers = df_all_outliers.sample(frac=1).reset_index(drop=True)

    # pad with zero-filled rows to make the number of rows divisible by 54
    # (the number of rows in a 9x6 grid)
    if df_all_outliers.shape[0] % 54 != 0:
        pad_df = pd.DataFrame(dtype=object).reindex_like(df_all_outliers)[0:54 - df_all_outliers.shape[0] % 54]
        # fill pad_df with zeros
        pad_df.fillna(0, inplace=True)
        df_all_outliers = pd.concat([df_all_outliers,
                                     pad_df],
                                    ignore_index=True, sort=False)

    # save to excel  file certain columns
    destination_excel_filename = data_folder + destination_run + f'{destination_run.split("/")[-2]}.xlsx'
    columns_for_outV_excel = ['reactions', 'DMF', 'ald001', 'ptsa', 'ptsa_dil_x_5', 'am001', 'ic001']
    df_all_outliers[columns_for_outV_excel].to_excel(destination_excel_filename,
                                                        index=False, sheet_name='reactions')

    # copy second sheet
    run_with_source_excel = 'multicomp-reactions/2023-06-20-run01/'
    path1 = data_folder + run_with_source_excel + f'{run_with_source_excel.split("/")[-2]}.xlsx'
    path2 = destination_excel_filename

    wb1 = xl.load_workbook(filename=path1)
    ws1 = wb1.worksheets[1]

    wb2 = xl.load_workbook(filename=path2)
    ws2 = wb2.create_sheet(ws1.title)

    for row in ws1:
        for cell in row:
            ws2[cell.coordinate].value = cell.value

    wb2.save(path2)

if __name__ == '__main__':
    # request_outliers(run_name='multicomp-reactions/2023-06-19-run01/',
    #                  destination_run='multicomp-reactions/2023-06-30-run01/',
    #                  known_outliers_filename='known_outliers',
    #                  manual_outliers_filename='manual_outliers',
    #                  n_repetitions_of_manual_outliers=2)

    request_outliers(run_name='multicomp-reactions/2023-06-19-run01/',
                     destination_run='multicomp-reactions/2023-07-04-run01/',
                     known_outliers_filename=False,
                     manual_outliers_filename='manual_outliers_2',
                     n_repetitions_of_manual_outliers=3)
