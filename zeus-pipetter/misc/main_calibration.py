"""
workflow:
1. initiate hardware
2. generate event list for detecting surface height of stock solutions
3. run events for surface detection, get liquid surface heights and write to excel
4. generate event list for pipetting
"""

import logging
import prepare_reaction as prep


def setup_logger():
    # better logging format in console
    class CustomFormatter(logging.Formatter):
        grey = "\x1b[38;20m"
        yellow = "\x1b[33;20m"
        red = "\x1b[31;20m"
        bold_red = "\x1b[31;1m"
        reset = "\x1b[0m"
        format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s (%(filename)s:%(lineno)d)"

        FORMATS = {
            logging.DEBUG: grey + format + reset,
            logging.INFO: yellow + format + reset,
            logging.WARNING: yellow + format + reset,
            logging.ERROR: red + format + reset,
            logging.CRITICAL: bold_red + format + reset
        }

        def format(self, record):
            log_fmt = self.FORMATS.get(record.levelno)
            formatter = logging.Formatter(log_fmt)
            return formatter.format(record)

    # create logger with 'main'
    logger = logging.getLogger('main')
    logger.setLevel(logging.DEBUG)
    # create file handler which logs even debug messages
    fh = logging.FileHandler('C:\\Yankai\\Dropbox\\robochem\\pipetter_files\\main.log')
    fh.setLevel(logging.INFO)
    # create console handler with a higher log level
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    # create formatter and add it to the handlers
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    ch.setFormatter(CustomFormatter())
    # add the handlers to the logger
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


logger = setup_logger()

import copy, time, pickle, re, importlib, json, os
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl import load_workbook
import PySimpleGUI as sg
import pandas as pd
# import arrow
import zeus
import pipetter
import planner as pln
import breadboard as brb

# data_folder = os.environ['ROBOCHEM_DATA_PATH'].replace('\\', '/') + '/'
# data_folder  = 'C:/Users/Chemiluminescence/Dropbox/robochem/data/'
#
data_folder = "C:\\Yankai\\Dropbox\\robochem\\pipetter_files"
def initiate_hardware() -> (zeus.ZeusModule, pipetter.Gantry, pipetter.Pipetter):
    # initiate zeus
    zm = zeus.ZeusModule(id=1)
    time.sleep(3)
    logger.info("zeus is loaded as: zm")

    # initiate gantry
    gt = pipetter.Gantry(zeus=zm)
    time.sleep(3)
    logger.info("gantry is loaded as: gt")
    # gt.configure_grbl() # This only need to be done once.
    gt.home_xy()
    if gt.xy_position == (0, 0):
        logger.info("gantry is homed")

    # initiate pipetter
    pt = pipetter.Pipetter(zeus=zm, gantry=gt, is_balance_involved = True)
    time.sleep(2)
    pt.close_balance_door()
    logger.info("pipetter is loaded as: pt")

    return zm, gt, pt

# use GUI to specify Excel path for reactions and stock solutions
def load_excel_path_by_pysimplegui():
    sg.theme('BrightColors')  # Add a touch of color
    working_directory = os.getcwd()
    # All the stuff inside your window.
    layout = [[sg.Text('Select Excel file for reactions')],
              [sg.InputText(key="-FILE_PATH-"),
               sg.FileBrowse(initial_folder=working_directory,
                             file_types=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))],
              [sg.Submit(), sg.Cancel()]]

    # Create the Window
    window = sg.Window('Select Excel file for reactions',
                       layout,
                       size=(600, 200),
                       font=('Helvetica', 14), )

    # Event Loop to process "events" and get the "values" of the inputs
    event, values = window.read()
    # print(event, values[0])
    window.close()
    logger.info(f"Excel file for reactions is selected: {values['-FILE_PATH-']}")
    return values['-FILE_PATH-']

def load_stock_solutions_from_excel(path: str) -> list:
    stock_solution_list = []
    wb_excel = load_workbook(path, data_only=True)
    ws = wb_excel[[x for x in wb_excel.sheetnames if 'stock_solutions' in x][0]]

    for row in tuple(ws.rows)[1:]:  # exclude the header
        if row[0].value is not None:
            substance_name = row[0].value,
            index = row[1].value,
            plate_id = row[2].value,
            container_id = row[3].value,
            solvent = row[4].value,
            density = row[5].value,
            volume = row[6].value,
            liquid_surface_height = row[7].value,
            pipetting_mode = row[8].value,
            stock_solution_list.append(
                {'substance_name': substance_name[0], 'index': index[0],
                 'plate_id': plate_id[0], 'container_id': container_id[0],
                 'solvent': solvent[0], 'density':density[0], 'volume': volume[0],
                 'liquid_surface_height': liquid_surface_height[0], 'pipetting_mode': pipetting_mode[0]})

    logger.info(f"stock solutions are loaded from Excel file: {stock_solution_list}")
    print(stock_solution_list)

    # stock_solution_list example:
    # {'substance_name': 'DMF', 'index': 'Substance_A', 'plate_id': 5, 'container_id': 0,
    # 'solvent': 'DMF', 'density': 0.944, 'volume': None, 'liquid_surface_height': 1804, 'pipetting_mode': 'empty'}

    return stock_solution_list

def update_stock_solution_list_to_excel(path_for_reactions: str, stock_solution_list: list):
    wb_excel = load_workbook(path_for_reactions)
    ws = wb_excel[[x for x in wb_excel.sheetnames if 'stock_solutions' in x][0]]
    for row in tuple(ws.rows)[1:]:  # exclude the header
        if row[0].value is not None:
            for solution in stock_solution_list:
                if row[0].value == solution['substance_name']:
                    row[6].value = solution['volume']
                    row[7].value = solution['liquid_surface_height']
    wb_excel.save(path_for_reactions)

def add_stock_solutions_to_containers(stock_solution_list: list, lld_type: str = "clld") -> list:
    containers_for_stock = []
    if lld_type == 'clld':
        liquid_class = 1
    elif lld_type == 'plld':
        liquid_class = 12
    for solution in stock_solution_list:
        solution_container = brb.plate_list[solution['plate_id']].containers[solution['container_id']]
        solution_container.substance = solution['substance_name']
        solution_container.substance_density = solution['density']
        solution_container.solvent = solution['solvent']

        solution_container.liquid_surface_height, solution_container.liquid_volume\
            = pt.check_volume_in_container(container=solution_container,liquidClassTableIndex=liquid_class,change_tip_after_each_check=True,
                                           tip_for_volume_check = '50ul')

        containers_for_stock.append(solution_container)

    logger.info(f"Stock solutions are added to containers: {containers_for_stock}")

    return containers_for_stock

    ## safety check

def check_if_event_list_legit(event_list: list):
        for event in event_list:
            assert event.asp_liquidClassTableIndex is not None, f"asp_liquidClassTableIndex is not correct: {event.asp_liquidClassTableIndex}"
            assert event.aspirationVolume >= 0, f"aspirationVolume is not correct: {event.aspirationVolume}"
            assert event.tip_type in ['50ul', '300ul', '1000ul'], f"tip type is not correct: {event.tip_type}"
            assert event.disp_liquidClassTableIndex is not None, f"disp_liquidClassTableIndex is not correct: {event.disp_liquidClassTableIndex}"
            assert event.dispenseVolume >= 0, f"dispenseVolume is not correct: {event.dispenseVolume}"

        print("event_list is legit!")

if __name__ == '__main__':
    #
    # ## initiate hardware
    # zm, gt, pt = initiate_hardware()
    #
    # path_for_reactions = load_excel_path_by_pysimplegui()
    # #
    # stock_solution_list = load_stock_solutions_from_excel(path=path_for_reactions)
    # # check the volume in stock containers and add stock solutions to containers
    # containers_for_stock = add_stock_solutions_to_containers(stock_solution_list, lld_type = "plld")
    # #
    # # # update the stock solution list
    # for solution in stock_solution_list:
    #     for container in containers_for_stock:
    #         if solution['substance_name'] == container.substance:
    #             solution['volume'] = container.liquid_volume
    #             solution['liquid_surface_height'] = container.liquid_surface_height
    # #
    # # update the stock solution list to Excel file
    # update_stock_solution_list_to_excel(path_for_reactions=path_for_reactions, stock_solution_list=stock_solution_list)
    #
    # df_reactions_grouped_by_plate_id,  substance_addition_sequence = prep.extract_reactions_df_to_run(path_for_reactions)
    #
    #
    # # generate event list for calibration
    # event_dataframe_chem, calibration_event_list = \
    #     pln.generate_event_object(logger=logger,
    #                               excel_to_generate_dataframe=path_for_reactions,
    #                               is_pipeting_to_balance=True, is_for_bio=False,
    #                               containers_for_stock=containers_for_stock)
    #
    #
    # pln.generate_event_list_new(df_reactions_grouped_by_plate_id = ,
    #                             substance_addition_sequence = [stock_solution_list[0]['substance_name']],
    #                             stock_solution_containers = containers_for_stock,
    #                             excel_path_for_conditions = path_for_reactions,
    #                             asp_lld = 0,
    #                             pipetting_to_balance=False)
    # time.sleep(1)

#

    ## initiate hardware
    zm, gt, pt = initiate_hardware()

    excel_path_before_treatment, \
    plate_barcodes, \
    reaction_temperature,\
    plate_barcode_for_dilution = prep.GUI_get_excel_path_plate_barcodes_temperature_etc()
    logger.info(f"excel_path_before_treatment: {excel_path_before_treatment}\n" \
    f"plate_barcodes: {plate_barcodes}\n" \
    f"reaction_temperature: {reaction_temperature}\n" \
    f"plate_barcode_for_dilution: {plate_barcode_for_dilution}")

    excel_path_for_conditions, _ = prep.prepare_excel_file_for_reaction(reaction_temperature=reaction_temperature,
                                                                        excel_path=excel_path_before_treatment,
                                                                        plate_barcodes=plate_barcodes,
                                                                        plate_barcodes_for_dilution=plate_barcode_for_dilution)
    # is_check_volume = prep.GUI_choose_if_check_surface_height()
    stock_solution_containers = \
        pln.assign_stock_solutions_to_containers_and_check_volume(excel_path = excel_path_for_conditions,
                                                                  check_volume_by_pipetter = True, pt=pt)

    df_reactions_grouped_by_plate_id,  substance_addition_sequence = prep.extract_reactions_df_to_run(excel_path_for_conditions)


    event_list_to_run = pln.generate_event_list_new(excel_path_for_conditions = excel_path_for_conditions,
                        df_reactions_grouped_by_plate_id = df_reactions_grouped_by_plate_id,
                        substance_addition_sequence = substance_addition_sequence,
                        stock_solution_containers = stock_solution_containers,
                        asp_lld = 0, pipetting_to_balance= True)

    for event in event_list_to_run:
        print(event.liquidClassTableIndex )
        event.liquidClassTableIndex = 37
        event.destination_container.liquid_surface_height = 1300
        # event.source_container.liquid_surface_height += 50

    # weighing_result = pln.do_calibration_on_events(zm=zm, pt=pt, logger=logger,
    #                                                calibration_event_list= event_list_to_run,
    #                                                change_tip_after_every_pipetting= False,
    #                                                repeat_n_times= 5, starting_index= 0)

# #########################################################################
# # specify tip and liquidClassIndex and other staff for calibration
# def specify_tip_and_liquidClassIndex_for_calibration(event_list: list = calibration_event_list):
#     for event in event_list:
#         # if event.substance_name == 'DMF_300ul':
#         #     event.tip_type = '300ul'
#         #     event.asp_liquidClassTableIndex = 22
#         #     event.disp_liquidClassTableIndex = 22
#         #     event.disp_liquidSurface = 1600
#         #     event.disp_lldSearchPosition = 1600
#         # elif event.substance_name == 'DMF_50ul':
#         #     event.tip_type = '50ul'
#         #     event.asp_liquidClassTableIndex = 24
#         #     event.disp_liquidClassTableIndex = 24
#         #     event.disp_liquidSurface = 1600
#         #     event.disp_lldSearchPosition = 1600
#         # elif event.substance_name == 'DMF_1000ul':
#         #     event.tip_type = '1000ul'
#         #     event.asp_liquidClassTableIndex = 23
#         #     event.disp_liquidClassTableIndex = 23
#         #     event.disp_liquidSurface = 1600
#         #     event.disp_lldSearchPosition = 1600
#         event.asp_liquidClassTableIndex = 1
#         event.disp_liquidClassTableIndex = 1
#         event.asp_lld = 0
#         event.asp_liquidSurface = 2000
#         event.disp_liquidSurface = 1700
#         event.aspirationVolume = 200
#         event.dispensingVolume = 200
#         event.tip_type = '300ul'
#
#
#     return event_list
#
#
# # calibration_event_list_adjust = specify_tip_and_liquidClassIndex_for_calibration()
# #########################################################################
#
#
#
# #############the below is for checking pipetting volume in 54 vials#################
# # calibration_event_list_reversed = calibration_event_list[::-1] # reverse the list. pipetting from large volume
#
# # for event in calibration_event_list_reversed:
# #     event.asp_liquidSurface = 1800
# #     event.asp_lldSearchPosition = 1800
# #     print(event)
# #
# # calib_list = [calibration_event_list[i] for i in [0]*54 ]
# # calib_list_run = []
# #
# # for i,event in enumerate(calib_list):
# #     event.asp_liquidSurface = 1800
# #     event.asp_lldSearchPosition = 1800
# #     event.source_container = brb.plate1.containers[i]
# #     calib_list_run.append(copy.deepcopy(event))
# #
# # calib_list_run = [[calib_list_run]]
# ##########################################
# for event in calibration_event_list:
#     event.asp_lld = 0
#     event.tip_type = '1000ul'
#     event.asp_liquidClassTableIndex = 29
#     event.disp_liquidClassTableIndex = 29
#
# # calibration_event_list = [[calibration_event_list]]
# # do_calibration

# # weighing_result = pln.do_calibration_on_events(zm=zm, pt=pt, logger=logger,
# #                                                calibration_event_list= calibration_event_list,
# #                                                change_tip_after_every_pipetting= False,
# #                                                repeat_n_times= 5, starting_index= 0)
